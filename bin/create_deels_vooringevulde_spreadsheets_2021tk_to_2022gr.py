#!/usr/bin/env python3

import csv
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pyexcel_ods3 import get_data, save_data

sys.path.insert(0, '.')


# This script takes a csv with the stembureaus of a previous election
# as input and copies/rewrites part of that data to create the
# 'deels vooringevulde spreadsheets' for each of the gemeenten of the
# upcoming elections


def as_text(value):
    if value is None:
        return ""
    return str(value)


# Names of fields that we want to copy from the stembureaus of the
# previous elections to the new elections
old_fieldnames = [
    'Nummer stembureau of afgiftepunt',
    'Naam stembureau of afgiftepunt',
    'Website locatie',
    'BAG referentienummer',
    'Extra adresaanduiding',
    'X',
    'Y',
    'Latitude',
    'Longitude',
    'Openingstijden 15-03-2021',
    'Openingstijden 16-03-2021',
    'Openingstijden 17-03-2021',
    'Mindervaliden toegankelijk',
    'Akoestiek',
    'Auditieve hulpmiddelen',
    'Visuele hulpmiddelen',
    'Mindervalide toilet aanwezig',
    'Tellocatie',
    'Contactgegevens',
    'Beschikbaarheid'
    #'Verkiezingen'
]


new_fieldnames = [
    'Nummer stembureau',
    'Naam stembureau',
    'Website locatie',
    'BAG Nummeraanduiding ID',
    'Extra adresaanduiding',
    'X',
    'Y',
    'Latitude',
    'Longitude',
    'Openingstijden 14-03-2022',
    'Openingstijden 15-03-2022',
    'Openingstijden 16-03-2022',
    'Toegankelijk voor mensen met een lichamelijke beperking',
    'Akoestiek',
    'Auditieve hulpmiddelen',
    'Visuele hulpmiddelen',
    'Gehandicaptentoilet',
    'Tellocatie',
    'Contactgegevens gemeente',
    'Verkiezingswebsite gemeente'
    #'Verkiezingen'
]


# Dict of gemeenten that have merged since the previous elections;
# we use this to combine the stembureaus of the old gemeenten into the
# deels vooringevulde spreadsheet of the newly formed gemeente
gemeente_herindelingen = {
    'Beemster': 'Purmerend',
    'Heerhugowaard': 'Dijk en Waard',
    'Langedijk': 'Dijk en Waard',
    'Landerd': 'Maashorst',
    'Uden': 'Maashorst',
    'Boxmeer': 'Land van Cuijk',
    'Cuijk': 'Land van Cuijk',
    'Grave': 'Land van Cuijk',
    'Mill en Sint Hubert': 'Land van Cuijk',
    'Sint Anthonis': 'Land van Cuijk'
}


def main():
    # CSV file from previous elections (download from CKAN)
    with open('files/eb2c1546-7f8d-41d4-9719-61b53b6d2111.csv') as IN:
        reader = csv.reader(IN)
        header = reader.__next__()
        records_per_gemeente = {}
        # Load the records
        for row in reader:
            data = dict(zip(header, row))

            # We don't use Afgiftepunten anymore in GR2022 as there is
            # no briefstemmen, so skip those rows
            if data['Stembureau of Afgiftepunt'] == 'Afgiftepunt':
                continue

            record = {}
            for key, value in data.items():
                if key in old_fieldnames:
                    # One time change to remove afgiftepunten from column name
                    if key == 'Nummer stembureau of afgiftepunt':
                        record['Nummer stembureau'] = value
                        continue
                    # One time change to remove afgiftepunten from column name
                    if key == 'Naam stembureau of afgiftepunt':
                        record['Naam stembureau'] = value
                        continue
                    # One time change of BAG column name
                    if key == 'BAG referentienummer':
                        record['BAG Nummeraanduiding ID'] = value
                        continue
                    # Use only the last 3 days as for GR2022 we only have 3
                    # days of elections instead of 8, and adjust the date
                    if key == 'Openingstijden 15-03-2021':
                        record['Openingstijden 14-03-2022'] = re.sub('20..-..-..', '2022-03-14', value)
                        continue
                    if key == 'Openingstijden 16-03-2021':
                        record['Openingstijden 15-03-2022'] = re.sub('20..-..-..', '2022-03-15', value)
                        continue
                    if key == 'Openingstijden 17-03-2021':
                        record['Openingstijden 16-03-2022'] = re.sub('20..-..-..', '2022-03-16', value)
                        continue
                    # One time change of column name and Y/N change
                    if key == 'Mindervaliden toegankelijk':
                        val = ''
                        if value == 'Y':
                            val = 'ja'
                        if value == 'N':
                            val = 'nee'
                        record['Toegankelijk voor mensen met een lichamelijke beperking'] = val
                        continue
                    # One time Y/N change
                    if key == 'Akoestiek':
                        val = ''
                        if value == 'Y':
                            val = 'ja'
                        if value == 'N':
                            val = 'nee'
                        record['Akoestiek'] = val
                        continue
                    # One time change of column name and Y/N change
                    if key == 'Mindervalide toilet aanwezig':
                        val = ''
                        if value == 'Y':
                            val = 'ja'
                        if value == 'N':
                            val = 'nee'
                        record['Gehandicaptentoilet'] = val
                        continue
                    # One time Y/N change
                    if key == 'Tellocatie':
                        val = ''
                        if value == 'Y':
                            val = 'ja'
                        if value == 'N':
                            val = 'nee'
                        record['Tellocatie'] = val
                        continue
                    # One time change of column name
                    if key == 'Contactgegevens':
                        record['Contactgegevens gemeente'] = value
                        continue
                    # One time change of column name
                    if key == 'Beschikbaarheid':
                        record['Verkiezingswebsite gemeente'] = value
                        continue
                    else:
                        record[key] = value

            # Combine stembureaus of gemeenten that have merged since
            # the previous elections
            gemeente = data['Gemeente']
            if gemeente in gemeente_herindelingen:
                gemeente = gemeente_herindelingen[gemeente]

            # Add this record to a gemeente (and create it if the
            # gemeente doesn't yet exist in the dict)
            try:
                records_per_gemeente[gemeente].append(record)
            except KeyError:
                records_per_gemeente[gemeente] = [record]

        if not os.path.exists('files/deels_vooringevuld'):
            os.mkdir('files/deels_vooringevuld')

        for gemeente, records in records_per_gemeente.items():
            # Save the records to the .xslx file
            print(gemeente)
            path = (
                "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
                "invulformulier_%s_deels_vooringevuld.xlsx" % (gemeente)
            )
            # Make sure to have an empty spreadsheet which is up to date
            # for the upcoming elections; copy this for each new gemeente
            # to create a deels vooringvulde spreadsheet
            os.system(
                "cp files/waarismijnstemlokaal.nl_invulformulier.xlsx "
                "\"%s\"" % (path)
            )
            workbook = load_workbook(path)
            worksheet = workbook['Attributen']
            font = Font(name="Arial", size=10)
            column = 6

            # Create mapping of row number and field name
            field_mapping = {}
            for idx, row in enumerate(range(1, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=1).value
                for new_fieldname in new_fieldnames:
                    if new_fieldname in cell_value:
                        field_mapping[new_fieldname] = idx + 1

            # Sort the records based on 'Nummer stembureau' while
            # placing records without a 'Nummer stembureau' at the end
            sorted_records = sorted(
                records,
                key=lambda k: int(
                    k['Nummer stembureau']
                ) if k['Nummer stembureau'] else 100000
            )

            # Save a record to a column
            for record in sorted_records:
                for new_fieldname in new_fieldnames:
                    worksheet.cell(
                        row=field_mapping[new_fieldname], column=column, value=record[new_fieldname]
                    ).font = font
                column += 1

            # Set width of each column
            for column_cells in tuple(worksheet.columns)[5:]:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = 52

            workbook.save(path)
            workbook.close()

            # Save the records to the .ods file
            #path = (
            #    "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
            #    "invulformulier_%s_deels_vooringevuld.ods" % (gemeente)
            #)
            #os.system(
            #    "cp files/waarismijnstemlokaal.nl_invulformulier.ods "
            #    "\"%s\"" % (path)
            #)
            #data = get_data(path)
            #for record in records:
            ## TODO this needs to be rewritten
            ##    data['Attributen'][2].append(record['Naam stembureau'])
            ##    if 'BAG referentienummer' in record:
            ##        data['Attributen'][5].append(
            ##            record['BAG referentienummer']
            ##        )
            ##    if 'Latitude' in record:
            ##        data['Attributen'][8].append(record['Latitude'])
            ##    if 'Longitude' in record:
            ##        data['Attributen'][7].append(record['Longitude'])
            ##    data['Attributen'][10].append(record['Openingstijden'])
            #    save_data(path, data)
    return 0

if __name__ == '__main__':
    sys.exit(main())
