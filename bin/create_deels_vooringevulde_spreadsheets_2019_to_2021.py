#!/usr/bin/env python3

import csv
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pyexcel_ods3 import get_data, save_data

sys.path.insert(0, '.')


def as_text(value):
    if value is None:
        return ""
    return str(value)

field_names = [
    'Nummer stembureau',
    'Naam stembureau',
    'Website locatie',
    'BAG referentienummer',
    'Extra adresaanduiding',
    'X',
    'Y',
    'Longitude',
    'Latitude',
    'Openingstijden',
    'Mindervaliden toegankelijk',
    'Akoestiek',
    'Mindervalide toilet aanwezig',
    'Contactgegevens',
    'Beschikbaarheid'
    #'Verkiezingen'
]

gemeente_herindelingen = {
    'Appingedam': 'Eemsdelta',
    'Delfzijl': 'Eemsdelta',
    'Loppersum': 'Eemsdelta',
    'Haaren': 'Oisterwijk',
    'Haaren': 'Vught',
    'Haaren': 'Boxtel',
    'Haaren': 'Tilburg'
}


def main():
    with open('files/b0e083ee-b44c-4573-9757-b92159087812.csv') as IN:
        reader = csv.reader(IN)
        header = reader.__next__()
        records_per_gemeente = {}
        # Load the records
        for row in reader:
            data = dict(zip(header, row))
            record = {}
            for key, value in data.items():
                if key in field_names:
                    if 'Openingstijden' in key:
                        record[key] = re.sub('20..-..-..', '2021-03-17', value)
                    else:
                        record[key] = value

            gemeente = data['Gemeente']
            if gemeente in gemeente_herindelingen:
                gemeente = gemeente_herindelingen[gemeente]

            try:
                records_per_gemeente[gemeente].append(record)
            except KeyError:
                records_per_gemeente[gemeente] = [record]

        if not os.path.exists('files/deels_vooringevuld'):
            os.mkdir('files/deels_vooringevuld')

        for gemeente, records in records_per_gemeente.items():
            # Save the records to the .xslx file
            print(gemeente)
            path = (
                "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
                "invulformulier_%s_deels_vooringevuld.xlsx" % (gemeente)
            )
            os.system(
                "cp files/waarismijnstemlokaal.nl_invulformulier.xlsx "
                "\"%s\"" % (path)
            )
            workbook = load_workbook(path)
            worksheet = workbook['Attributen']
            font = Font(name="Arial", size=10)
            column = 6

            # Create mapping of row number and field name
            field_mapping = {}
            for idx, row in enumerate(range(1, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=1).value
                for field_name in field_names:
                    orig_field_name = field_name
                    if field_name == 'BAG referentienummer':
                        field_name = 'BAG referentie nummer'
                    if field_name in cell_value:
                        field_mapping[orig_field_name] = idx + 1

            # Sort the records based on 'Nummer stembureau' while
            # placing records without a 'Nummer stembureau' at the end
            sorted_records = sorted(
                records,
                key=lambda k: int(
                    k['Nummer stembureau']
                ) if k['Nummer stembureau'] else 100000
            )

            # Save a record to a column
            for record in sorted_records:
                for field_name in field_names:
                    worksheet.cell(
                        row=field_mapping[field_name], column=column, value=record[field_name]
                    ).font = font
                worksheet.cell(
                    row=2, column=column, value='Stembureau'
                ).font = font
                column += 1

            # Set width of each column
            for column_cells in tuple(worksheet.columns)[5:]:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = 52

            workbook.save(path)
            workbook.close()

            # Save the records to the .ods file
            #path = (
            #    "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
            #    "invulformulier_%s_deels_vooringevuld.ods" % (gemeente)
            #)
            #os.system(
            #    "cp files/waarismijnstemlokaal.nl_invulformulier.ods "
            #    "\"%s\"" % (path)
            #)
            #data = get_data(path)
            #for record in records:
            ## TODO this needs to be rewritten
            ##    data['Attributen'][2].append(record['Naam stembureau'])
            ##    if 'BAG referentienummer' in record:
            ##        data['Attributen'][5].append(
            ##            record['BAG referentienummer']
            ##        )
            ##    if 'Longitude' in record:
            ##        data['Attributen'][7].append(record['Longitude'])
            ##    if 'Latitude' in record:
            ##        data['Attributen'][8].append(record['Latitude'])
            ##    data['Attributen'][10].append(record['Openingstijden'])
            #    save_data(path, data)
    return 0

if __name__ == '__main__':
    sys.exit(main())
