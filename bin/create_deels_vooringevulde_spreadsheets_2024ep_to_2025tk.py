#!/usr/bin/env python3

import csv
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pyexcel_ods3 import get_data, save_data

sys.path.insert(0, '.')


# This script takes a csv with the stembureaus of a previous election
# as input and copies/rewrites part of that data to create the
# 'deels vooringevulde spreadsheets' for each of the gemeenten of the
# upcoming elections


def as_text(value):
    if value is None:
        return ""
    return str(value)


# Names of fields that we want to copy from the stembureaus of the
# previous elections to the new elections
old_fieldnames = [
    'Nummer stembureau',
    'Naam stembureau',
    'Type stembureau',
    'Website locatie',
    'BAG Nummeraanduiding ID',
    'Extra adresaanduiding',
    # X en Y weglaten, want Lat/Long zijn toch leidend en dit geeft problemen bij de BES-eilanden?
    #'X',
    #'Y',
    'Latitude',
    'Longitude',
    'Openingstijd',
    'Sluitingstijd',
    'Toegankelijk voor mensen met een lichamelijke beperking',
    'Toegankelijke ov-halte',
    'Gehandicaptentoilet',
    'Host',
    'Geleidelijnen',
    'Stemmal met audio-ondersteuning',
    'Kandidatenlijst in braille',
    'Kandidatenlijst met grote letters',
    'Gebarentolk (NGT)',
    'Gebarentalig stembureaulid (NGT)',
    'Akoestiek geschikt voor slechthorenden',
    'Prikkelarm',
    'Extra toegankelijkheidsinformatie',
    'Overige informatie',
    'Tellocatie',
    'Contactgegevens gemeente',
    'Verkiezingswebsite gemeente'
    #'Verkiezingen'
]


new_fieldnames = [
    'Nummer stembureau',
    'Naam stembureau',
    'Type stembureau',
    'Website locatie',
    'BAG Nummeraanduiding ID',
    'Extra adresaanduiding',
    # X en Y weglaten, want Lat/Long zijn toch leidend en dit geeft problemen bij de BES-eilanden?
    #'X',
    #'Y',
    'Latitude',
    'Longitude',
    'Openingstijd',
    'Sluitingstijd',
    'Toegankelijk voor mensen met een lichamelijke beperking',
    'Toegankelijke ov-halte',
    'Gehandicaptentoilet',
    'Host',
    'Geleidelijnen',
    'Stemmal met audio-ondersteuning',
    'Kandidatenlijst in braille',
    'Kandidatenlijst met grote letters',
    'Gebarentolk (NGT)',
    'Gebarentalig stembureaulid (NGT)',
    'Akoestiek geschikt voor slechthorenden',
    'Prikkelarm',
    'Extra toegankelijkheidsinformatie',
    'Overige informatie',
    'Tellocatie',
    'Contactgegevens gemeente',
    'Verkiezingswebsite gemeente',
    #'Verkiezingen'
]


# Dict of gemeenten that have merged since the previous elections;
# we use this to combine the stembureaus of the old gemeenten into the
# deels vooringevulde spreadsheet of the newly formed gemeente
#gemeente_herindelingen = {
#    'Brielle': 'Voorne aan Zee',
#    'Hellevoetsluis': 'Voorne aan Zee',
#    'Westvoorne': 'Voorne aan Zee'
#}


def main():
    # CSV file from previous elections (download from CKAN)
    # 2024EP
    with open('files/b6061c88-3736-4ab4-bcbf-ba8e980e0fe7.csv') as IN:
        reader = csv.reader(IN)
        header = reader.__next__()
        records_per_gemeente = {}
        # Load the records
        for row in reader:
            data = dict(zip(header, row))

            # There is only 1 elections day for 2023PS instead of 3 for 2022GR.
            # We only want to copy stembureaus that were open on Wednesday,
            # skip the rest.
            #if not data['Openingstijden 16-03-2022']:
            #    continue

            record = {}

            # If there are new fields this election, then add them here with
            # there default value
            #record['Verkiezingen'] = ''
            record['Host'] = ''
            record['Geleidelijnen'] = ''
            record['Stemmal met audio-ondersteuning'] = ''
            record['Kandidatenlijst in braille'] = ''
            record['Kandidatenlijst met grote letters'] = ''
            record['Gebarentolk (NGT)'] = ''
            record['Gebarentalig stembureaulid (NGT)'] = ''
            record['Prikkelarm'] = ''
            record['Overige informatie'] = ''
            for key, value in data.items():
                if key in old_fieldnames:
                    # One time change of column value,  because we changed it
                    # from free text to ja/nee
                    if key == 'Toegankelijke ov-halte':
                        if value == 'n.v.t.':
                            record['Toegankelijke ov-halte'] = ''
                        elif value:
                            record['Toegankelijke ov-halte'] = 'ja'
                        else:
                            record['Toegankelijke ov-halte'] = ''
                        continue
                    # One time fix because of importing data from json -> csv
                    #backup which resulted in some null instead of empty values
                    if key == 'Website locatie':
                        if value == 'null':
                            record['Website locatie'] = ''
                        else:
                            record['Website locatie'] = value
                        continue
                    # Clear the value of extra toegankelijkheidsinformatie for
                    # only this election because the previous election also
                    # contains TSA kenmerken
                    elif key == 'Extra toegankelijkheidsinformatie':
                            record['Extra toegankelijkheidsinformatie'] = ''
                    # Change openingstijd
                    elif key == 'Openingstijd':
                        record['Openingstijd'] = re.sub('20..-..-..', '2025-10-29', value)
                        continue
                    # Change sluitingstijd
                    elif key == 'Sluitingstijd':
                        record['Sluitingstijd'] = re.sub('20..-..-..', '2025-10-29', value)
                        continue
                    else:
                        record[key] = value

            # Combine stembureaus of gemeenten that have merged since
            # the previous elections
            gemeente = data['Gemeente']
            #if gemeente in gemeente_herindelingen:
            #    gemeente = gemeente_herindelingen[gemeente]

            # Add this record to a gemeente (and create it if the
            # gemeente doesn't yet exist in the dict)
            try:
                records_per_gemeente[gemeente].append(record)
            except KeyError:
                records_per_gemeente[gemeente] = [record]

        if not os.path.exists('files/deels_vooringevuld'):
            os.mkdir('files/deels_vooringevuld')

        for gemeente, records in records_per_gemeente.items():
            # Save the records to the .xslx file
            print(gemeente)
            path = (
                "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
                "invulformulier_%s_deels_vooringevuld.xlsx" % (gemeente)
            )
            # Make sure to have an empty spreadsheet which is up to date
            # for the upcoming elections; copy this for each new gemeente
            # to create a deels vooringvulde spreadsheet
            os.system(
                "cp files/waarismijnstemlokaal.nl_invulformulier.xlsx "
                "\"%s\"" % (path)
            )
            workbook = load_workbook(path)
            worksheet = workbook['Attributen']
            font = Font(name="Arial", size=10)
            column = 6

            # Create mapping of row number and field name
            field_mapping = {}
            for idx, row in enumerate(range(1, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=1).value
                for new_fieldname in new_fieldnames:
                    if new_fieldname in cell_value:
                        field_mapping[new_fieldname] = idx + 1


            # Sort the records based on 'Nummer stembureau' while
            # placing records without a 'Nummer stembureau' at the end
            sorted_records = sorted(
                records,
                key=lambda k: int(
                    k['Nummer stembureau']
                ) if k['Nummer stembureau'] else 100000
            )

            # Save a record to a column
            for record in sorted_records:
                for new_fieldname in new_fieldnames:
                    worksheet.cell(
                        row=field_mapping[new_fieldname], column=column, value=record[new_fieldname]
                    ).font = font
                column += 1

            # Set width of each column
            for column_cells in tuple(worksheet.columns)[5:]:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = 52

            workbook.save(path)
            workbook.close()

            # Save the records to the .ods file
            #path = (
            #    "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
            #    "invulformulier_%s_deels_vooringevuld.ods" % (gemeente)
            #)
            #os.system(
            #    "cp files/waarismijnstemlokaal.nl_invulformulier.ods "
            #    "\"%s\"" % (path)
            #)
            #data = get_data(path)
            #for record in records:
            ## TODO this needs to be rewritten
            ##    data['Attributen'][2].append(record['Naam stembureau'])
            ##    if 'BAG referentienummer' in record:
            ##        data['Attributen'][5].append(
            ##            record['BAG referentienummer']
            ##        )
            ##    if 'Latitude' in record:
            ##        data['Attributen'][8].append(record['Latitude'])
            ##    if 'Longitude' in record:
            ##        data['Attributen'][7].append(record['Longitude'])
            ##    data['Attributen'][10].append(record['Openingstijden'])
            #    save_data(path, data)
    return 0

if __name__ == '__main__':
    sys.exit(main())
