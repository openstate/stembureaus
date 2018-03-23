#!/usr/bin/env python

import os
import sys
import csv

from openpyxl import load_workbook
from openpyxl.styles import Font
from pyexcel_ods3 import get_data, save_data

sys.path.insert(0, '.')

from app.utils import find_buurt_and_wijk
from app.models import BAG


def find_bag_record(data):
    return BAG.query.filter_by(
        openbareruimte=data['Straat'].strip(),
        huisnummer=data['Huisnummer'].strip(),
        huisnummertoevoeging=data['Toevoeging'].strip(),
        woonplaats=data['Plaats'].strip()
    ).first()


def find_bag_record_by_huisletter(data):
    return BAG.query.filter_by(
        openbareruimte=data['Straat'].strip(),
        huisnummer=data['Huisnummer'].strip(),
        huisletter=data['Toevoeging'].strip(),
        woonplaats=data['Plaats'].strip()
    ).first()

def as_text(value):
    if value is None:
        return ""
    return str(value)

def main():
    reader = csv.reader(sys.stdin, delimiter=',')
    header = reader.__next__()
    points_per_muni = {}
    for row in reader:
        data = dict(zip(header, row))
        tijd_open, tijd_sluit = data['Openingsti'].split(' - ', 1)
        bag = find_bag_record(data)
        if bag is None:
            bag = find_bag_record_by_huisletter(data)
        if bag is None:
            record = {
                'Naam stembureau': data['Naam'],
                'Openingstijden': '2018-03-21T%s:00 tot 2018-03-21T%s:00' % (
                    tijd_open, tijd_sluit)
            }
        else:
            record = {
                'BAG referentienummer': bag.nummeraanduiding,
                'Naam stembureau': data['Naam'],
                'Openingstijden': '2018-03-21T%s:00 tot 2018-03-21T%s:00' % (
                    tijd_open, tijd_sluit),
                'Longitude': float(bag.lon),
                'Latitude': float(bag.lat)
            }
        try:
            points_per_muni[data['Gemeente']].append(record)
        except KeyError:
            points_per_muni[data['Gemeente']] = [record]

    for muni, points in points_per_muni.items():
        # first Excel
        print(muni)
        path = "data/waarismijnstemlokaal.nl_invulformulier_%s_deels_vooringevuld.xlsx" % (muni,)
        os.system("cp files/waarismijnstemlokaal.nl_invulformulier.xlsx \"%s\"" % (
            path,))
        workbook = load_workbook(path)
        worksheet = workbook['Attributen']
        font = Font(name="Arial", size=10)
        col = 6
        for point in points:
            worksheet.cell(row=3, column=col, value=point['Naam stembureau']).font = font

            if 'BAG referentienummer' in point:
                worksheet.cell(row=5, column=col, value=point['BAG referentienummer']).font = font
            if 'Longitude' in point:
                worksheet.cell(row=7, column=col, value=point['Longitude']).font = font
            if 'Latitude' in point:
                worksheet.cell(row=8, column=col, value=point['Latitude']).font = font
            worksheet.cell(row=10, column=col, value=point['Openingstijden']).font = font
            col += 1

        # Set width of each column
        for column_cells in tuple(worksheet.columns)[5:]:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column].width = int(length * 0.75)

        workbook.save(path)
        workbook.close()

        # now openoffice
        #path = "data/%s-2018-03-21.ods" % (muni,)
        #os.system("cp files/waarismijnstemlokaal.nl_invulformulier.ods \"%s\"" % (
        #    path,))
        #data = get_data(path)
        #for point in points:
        #    data['Attributen'][2].append(point['Naam stembureau'])
        #    if 'BAG referentienummer' in point:
        #        data['Attributen'][5].append(point['BAG referentienummer'])
        #    if 'Longitude' in point:
        #        data['Attributen'][7].append(point['Longitude'])
        #    if 'Latitude' in point:
        #        data['Attributen'][8].append(point['Latitude'])
        #    data['Attributen'][10].append(point['Openingstijden'])
        #    save_data(path, data)
    return 0

if __name__ == '__main__':
    sys.exit(main())
