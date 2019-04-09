#!/usr/bin/env python3

import csv
import os
import re
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font
from pyexcel_ods3 import get_data, save_data

sys.path.insert(0, '.')


def as_text(value):
    if value is None:
        return ""
    return str(value)

field_names = [
    'Nummer stembureau',
    'Naam stembureau',
    'Website locatie',
    'BAG referentienummer',
    'Extra adresaanduiding',
    'X',
    'Y',
    'Longitude',
    'Latitude',
    'Openingstijden',
    'Mindervaliden toegankelijk',
    'Akoestiek',
    'Mindervalide toilet aanwezig',
    'Contactgegevens',
    'Beschikbaarheid',
    #'Verkiezingen',
    "1.1.a Aanduiding aanwezig",
    "1.1.b Aanduiding duidelijk zichtbaar",
    "1.1.c Aanduiding goed leesbaar",
    "1.2.a GPA aanwezig",
    "1.2.b Aantal vrij parkeerplaatsen binnen 50m van de entree",
    "1.2.c Hoogteverschil tussen parkeren en trottoir",
    "1.2.d Hoogteverschil",
    "1.2.e Type overbrugging",
    "1.2.f Overbrugging conform ITstandaard",
    "1.3.a Vlak, verhard en vrij van obstakels",
    "1.3.b Hoogteverschil",
    "1.3.c Type overbrugging",
    "1.3.d Overbrugging conform ITstandaard",
    "1.3.e Obstakelvrije breedte van de route",
    "1.3.f Obstakelvrije hoogte van de route",
    "1.4.a Is er een route tussen gebouwentree en stemruimte",
    "1.4.b Route duidelijk aangegeven",
    "1.4.c Vlak en vrij van obstakels",
    "1.4.d Hoogteverschil",
    "1.4.e Type overbrugging",
    "1.4.f Overbrugging conform ITstandaard",
    "1.4.g Obstakelvrije breedte van de route",
    "1.4.h Obstakelvrije hoogte van de route",
    "1.4.i Deuren in route bedien- en bruikbaar",
    "2.1.a Deurtype",
    "2.1.b Opstelruimte aan beide zijden van de deur",
    "2.1.c Bedieningskracht buitendeur",
    "2.1.d Drempelhoogte (t.o.v. straat/vloer niveau)",
    "2.1.e Vrije doorgangsbreedte buitendeur",
    "2.2.a Tussendeuren aanwezig in eventuele route",
    "2.2.b Deurtype",
    "2.2.c Opstelruimte aan beide zijden van de deur",
    "2.2.d Bedieningskracht deuren",
    "2.2.e Drempelhoogte (t.o.v. vloer niveau)",
    "2.2.f Vrije doorgangsbreedte deur",
    "2.3.a Deur aanwezig naar/van stemruimte",
    "2.3.b Deurtype",
    "2.3.c Opstelruimte aan beide zijden van de deur",
    "2.3.d Bedieningskracht deur",
    "2.3.e Drempelhoogte (t.o.v. vloer niveau)",
    "2.3.f Vrije doorgangsbreedte deur",
    "2.4.a Zijn er tijdelijke voorzieningen aangebracht",
    "2.4.b VLOERBEDEKKING: Randen over de volle lengte deugdelijk a",
    "2.4.c HELLINGBAAN: Weerbestendig (alleen van toepassing bij bu",
    "2.4.d HELLINGBAAN: Deugdelijk verankerd aan ondergrond",
    "2.4.e LEUNING BIJ HELLINGBAAN/TRAP: Leuning aanwezig en confor",
    "2.4.f DORPELOVERBRUGGING: Weerbestendig (alleen van toepassing",
    "2.4.g DORPELOVERBRUGGING: Deugdelijk verankerd aan ondergrond",
    "3.1.a Obstakelvrije doorgangen",
    "3.1.b Vrije draaicirkel / manoeuvreerruimte",
    "3.1.c Idem voor stemtafel en stemhokje",
    "3.1.d Opstelruimte voor/naast stembus",
    "3.2.a Stoelen in stemruimte aanwezig",
    "3.2.b 1 op 5 Stoelen uitgevoerd met armleuningen",
    "3.3.a Hoogte van het laagste schrijfblad",
    "3.3.b Schrijfblad onderrijdbaar",
    "3.4.a Hoogte inworpgleuf stembiljet",
    "3.4.b Afstand inwerpgleuf t.o.v. de opstelruimte",
    "3.5.a Leesloep (zichtbaar) aanwezig",
    "3.6.a Kandidatenlijst in stemlokaal aanwezig",
    "3.6.b Opstelruimte voor de kandidatenlijst aanwezig"
]

gemeente_herindelingen = {
    'Haren': 'Groningen',
    'Ten Boer': 'Groningen',
    'Binnenmaas': 'Hoeksche Waard',
    'Cromstrijen': 'Hoeksche Waard',
    'Korendijk': 'Hoeksche Waard',
    'Oud-Beijerland': 'Hoeksche Waard',
    'Strijen': 'Hoeksche Waard',
    'Leerdam': 'Vijfheerenlanden',
    'Vianen': 'Vijfheerenlanden',
    'Zederik': 'Vijfheerenlanden',
    'Aalburg': 'Altena',
    'Werkendam': 'Altena',
    'Woudrichem': 'Altena',
    'Nuth': 'Beekdaelen',
    'Onderbanken': 'Beekdaelen',
    'Schinnen': 'Beekdaelen',
    'Haarlemmerliede en Spaarnwoude': 'Haarlemmermeer',
    'Bedum': 'Het Hogeland',
    'De Marne': 'Het Hogeland',
    'Eemsmond': 'Het Hogeland',
    'Winsum': 'Het Hogeland',
    'Grootegast': 'Westerkwartier',
    'Leek': 'Westerkwartier',
    'Marum': 'Westerkwartier',
    'Zuidhorn': 'Westerkwartier',
    'Giessenlanden': 'Molenlanden',
    'Molenwaard': 'Molenlanden',
    'Dongeradeel': 'Noardeast-Fryslân',
    'Ferwerderadiel': 'Noardeast-Fryslân',
    'Kollumerland en Nieuwkruisland': 'Noardeast-Fryslân',
    'Noordwijkerhout': 'Noordwijk',
    'Geldermalsen': 'West Betuwe',
    'Lingewaal': 'West Betuwe',
    'Neerijnen': 'West Betuwe'
}


def main():
    #with open('files/67f5d14c-b625-485f-9f47-8faccc5e27bc.csv') as IN:
    with open('files/b0e083ee-b44c-4573-9757-b92159087812.csv') as IN:
        reader = csv.reader(IN)
        header = reader.__next__()
        records_per_gemeente = {}
        # Load the records
        for row in reader:
            data = dict(zip(header, row))
            record = {}
            for key, value in data.items():
                if key in field_names:
                    #if key == 'Openingstijden':
                    #    record[key] = re.sub('2018-03-21', '2019-03-20', value)
                    #else:
                    #    record[key] = value
                    record[key] = value

            gemeente = data['Gemeente']
            if gemeente in gemeente_herindelingen:
                gemeente = gemeente_herindelingen[gemeente]

            try:
                records_per_gemeente[gemeente].append(record)
            except KeyError:
                records_per_gemeente[gemeente] = [record]

        if not os.path.exists('files/deels_vooringevuld'):
            os.mkdir('files/deels_vooringevuld')

        for gemeente, records in records_per_gemeente.items():
            # Save the records to the .xslx file
            print(gemeente)
            path = (
                "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
                "invulformulier_%s_deels_vooringevuld.xlsx" % (gemeente)
            )
            os.system(
                "cp files/waarismijnstemlokaal.nl_invulformulier.xlsx "
                "\"%s\"" % (path)
            )
            workbook = load_workbook(path)
            worksheet = workbook['Attributen']
            font = Font(name="Arial", size=10)
            column = 6

            # Retrieve mapping of row number and field name
            field_mapping = {}
            for idx, row in enumerate(range(1, worksheet.max_row + 1)):
                cell_value = worksheet.cell(row=row, column=1).value
                for field_name in field_names:
                    orig_field_name = field_name
                    if field_name == 'BAG referentienummer':
                        field_name = 'BAG referentie nummer'
                    if field_name in cell_value:
                        field_mapping[orig_field_name] = idx + 1

            # Sort the records based on 'Nummer stembureau' while
            # placing records without a 'Nummer stembureau' at the end
            sorted_records = sorted(
                records,
                key=lambda k: int(
                    k['Nummer stembureau']
                ) if k['Nummer stembureau'] else 100000
            )

            for record in sorted_records:
                for field_name in field_names:
                    worksheet.cell(
                        row=field_mapping[field_name], column=column, value=record[field_name]
                    ).font = font
                column += 1

            # Set width of each column
            for column_cells in tuple(worksheet.columns)[5:]:
                length = max(len(as_text(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[
                    column_cells[0].column
                ].width = int(length * 0.75)

            workbook.save(path)
            workbook.close()

            # Save the records to the .ods file
            #path = (
            #    "files/deels_vooringevuld/waarismijnstemlokaal.nl_"
            #    "invulformulier_%s_deels_vooringevuld.ods" % (gemeente)
            #)
            #os.system(
            #    "cp files/waarismijnstemlokaal.nl_invulformulier.ods "
            #    "\"%s\"" % (path)
            #)
            #data = get_data(path)
            #for record in records:
            ## TODO this needs to be rewritten
            ##    data['Attributen'][2].append(record['Naam stembureau'])
            ##    if 'BAG referentienummer' in record:
            ##        data['Attributen'][5].append(
            ##            record['BAG referentienummer']
            ##        )
            ##    if 'Longitude' in record:
            ##        data['Attributen'][7].append(record['Longitude'])
            ##    if 'Latitude' in record:
            ##        data['Attributen'][8].append(record['Latitude'])
            ##    data['Attributen'][10].append(record['Openingstijden'])
            #    save_data(path, data)
    return 0

if __name__ == '__main__':
    sys.exit(main())
